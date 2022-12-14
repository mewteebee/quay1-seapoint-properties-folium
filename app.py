import folium as fl
from openpyxl import load_workbook

def tuple_to_list(x):
    return [x[y].value for y in range(len(x))]

# loading excel file into vsc
wb = load_workbook('cpt-homes.xlsx')
ws = wb.worksheets[0]
property_name = ws['A']
latitude = ws['B']
longitude = ws['C']
pictures = ws['D']
links = ws['E']

# pulling data from columns into lists 
property_name_list = tuple_to_list(property_name)
latitude_list = tuple_to_list(latitude)
longitude_list = tuple_to_list(longitude)
pictures_list = tuple_to_list(pictures)
links_list = tuple_to_list(links)

# data checks
print(property_name_list)
print(latitude_list)
print(longitude_list)
print(pictures_list)
print(links_list)

# instantiating map object 
map = fl.Map(location=[-33.92714772961316, 18.41353385511342], zoom_start=13, tiles="Stamen Terrain")

# creating feature group 
fg = fl.FeatureGroup(name="myMap")

# placing markers at the property locations using data from imported xlsx file
for name, lat, lon, pic, link in zip(property_name_list, latitude_list, longitude_list, pictures_list, links_list):
    
    # creating html for the map markers
    html = f""" 
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.2.0-beta1/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-0evHe/X+R7YkIZDRvuzKMRqM+OrBnVFBL6DOitfPri4tjfHxaWutUpFmBp4vmVor" crossorigin="anonymous">
            </head>
            <div style="display:inline-block">
            <center>
                <h1 style="color:lightskyblue;"> {name} </h1>
            </center>
            <center>
                <img src="{pic}" style="height:auto; max-width:100%; margin:20px;" alt="Image of {name}"></img>
            </center>
            <center>
                <a href="{link}" type="button" target="_blank" rel="noopener" class="btn btn-light btn-block">Visit Page</a>            
            </center>
            </div>
            </html>
            """
    iframe = fl.IFrame(html, width = 250, height = 400)
    popup = fl.Popup(iframe)
    fg.add_child(fl.Marker(location = [lat, lon], popup = (popup) ))

map.add_child(fg)

map.save("index.html")